"""
Pipeline de limpieza y carga completa de datos PDM
desde 'PDM (1).xlsx' → SQL para Supabase.

Datos: 2024 (completo), 2025 (completo), 2026 (parcial), 2027 (sin ejecución).
Todos los cálculos ya están hechos en el Excel — se pasan tal cual, sin caps ni recálculos.
"""

import openpyxl
import json
import csv
from pathlib import Path

INPUT_FILE = "PDM (1).xlsx"
OUTPUT_SQL = "013_reload_pdm_may2026.sql"
OUTPUT_PATH = Path(__file__).parent.parent / "backend" / "src" / "db" / "migrations" / OUTPUT_SQL
VALID_YEARS = [2024, 2025, 2026, 2027]

# ── Column mapping (1-based Excel col index → DB field) ────────────────────
COL = {
    "meta_num": 1, "cod_dependencia": 2, "num_pilar": 3, "nom_pilar": 4,
    "macrometa": 5, "cod_sector": 6, "nom_sector": 7, "cod_programa": 8,
    "nombre_programa": 9, "cod_bpin": 10, "acciones_plan": 11,
    "unidad_medida": 12, "linea_base": 13, "anio_base": 14,
    "codigo_producto": 15, "indicador_meta": 16, "secretaria": 17,
    "descripcion_meta": 18, "meta_cuatrienio": 19, "tipo_ponderado": 20,

    "meta_pdm_2024": 21, "meta_fisica_2024": 22,
    "meta_pdm_2025": 23, "meta_fisica_2025": 24,
    "meta_pdm_2026": 25, "meta_fisica_2026": 26,
    "meta_pdm_2027": 27, "meta_fisica_2027": 28,

    # Presupuesto 2024
    "total_apropiacion_2024": 31, "total_apropiacion_imder_2024": 32,
    "neto_registros_2024": 33, "total_obligacion_2024": 34,
    "pct_ejec_obligado_2024": 36, "pct_ejec_pagos_2024": 37,
    "presup_comprometer_2024": 38,
    # Presupuesto 2025
    "total_apropiacion_2025": 41, "neto_registros_2025": 43,
    "total_obligacion_2025": 44, "pct_ejec_obligado_2025": 46,
    "pct_ejec_pagos_2025": 47, "presup_comprometer_2025": 48,
    # Presupuesto 2026
    "total_apropiacion_2026": 66, "total_apropiacion_imder_2026": 67,
    "neto_registros_2026": 68, "total_obligacion_2026": 69,
    "pct_ejec_obligado_2026": 71, "pct_ejec_pagos_2026": 72,
    "presup_comprometer_2026": 73,
    # Presupuesto 2027
    "total_apropiacion_2027": 91, "total_apropiacion_imder_2027": 92,
    "neto_registros_2027": 93, "total_obligacion_2027": 94,
    "pct_ejec_obligado_2027": 96, "pct_ejec_pagos_2027": 97,
    "presup_comprometer_2027": 98,

    # Calculated metrics (from Excel — used as-is)
    "cumplimiento_2024": 99, "avfis_2024": 100,
    "cumplimiento_2025": 101, "avfis_2025": 102,
    "cumplimiento_2026": 103, "avfis_2026": 104,
    "cumplimiento_2027": 105, "avfis_2027": 106,
    "cumplimiento_cuatrienio": 107, "avance_fisico": 108,

    # Obs / Compromisos
    "observaciones_2024": 109, "compromisos_2024": 110,
    "observaciones_2025": 111, "compromisos_2025": 112,
    "observaciones_2026": 113, "compromisos_2026": 114,
    "observaciones_2027": 115, "compromisos_2027": 116,

    # Eficiencia
    "eficiencia_2024": 117, "eficiencia_2025": 118,
    "eficiencia_2026": 119, "eficiencia_2027": 120,

    "avance_financiero": 121,
}


# ── Cleaning (format only, NO capping, NO recalculation) ───────────────────

def clean_currency(val):
    if val is None: return None
    s = str(val).strip()
    if s in ("", "-", "nan", "$-", "$ -", "None", "$0.00"): return None
    s = s.replace("$", "").replace(" ", "")
    neg = s.startswith("-")
    if neg: s = s[1:]
    if not s or s == "-": return None
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."): s = s.replace(".", "").replace(",", ".")
        else: s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        s = s.replace(",", ".") if len(parts[-1]) <= 2 else s.replace(",", "")
    elif "." in s and s.count(".") > 1:
        s = s.replace(".", "")
    try:
        v = float(s)
        return -v if neg else v
    except ValueError: return None


def clean_pct(val):
    if val is None: return None
    s = str(val).strip()
    if s.upper() in ("", "-", "NAN", "NP", "N.P", "NONE", "N/A"): return None
    has_pct = "%" in s
    s = s.replace("%", "").replace(",", ".").strip()
    try:
        v = float(s)
        return v / 100.0 if has_pct else v
    except ValueError: return None


def clean_num(val):
    if val is None: return None
    s = str(val).strip()
    if s.upper() in ("", "-", "NAN", "NP", "N.P", "NONE", "N/A"): return None
    s = s.replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."): s = s.replace(".", "").replace(",", ".")
        else: s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        s = s.replace(",", ".") if len(parts[-1]) <= 2 else s.replace(",", "")
    try: return float(s)
    except ValueError: return None


def clean_text(val):
    if val is None: return None
    s = str(val).strip()
    return None if s.upper() in ("", "NAN", "NONE") else s


def is_np(val):
    if val is None: return True
    return str(val).strip().upper() in ("NP", "N.P", "N/A", "")


# ── SQL helpers ────────────────────────────────────────────────────────────

def sql_val(val, is_text=False):
    if val is None: return "NULL"
    if is_text:
        return "'" + str(val).replace("'", "''").replace("\r\n", " ").replace("\n", " ") + "'"
    return str(val)

def sql_jsonb(obj):
    if not obj or all(v is None for v in obj.values()): return "NULL"
    clean = {k: v for k, v in obj.items() if v is not None}
    return "NULL" if not clean else f"'{json.dumps(clean)}'::jsonb"


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    print(f"[1] Loading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb.active
    print(f"    Sheet: {ws.title} | {ws.max_row} rows x {ws.max_column} cols")

    records = []
    skipped = 0

    for r in range(2, ws.max_row + 1):
        meta_raw = ws.cell(r, COL["meta_num"]).value
        if meta_raw is None:
            skipped += 1
            if skipped > 10: break
            continue
        skipped = 0

        try: meta_num = int(float(str(meta_raw)))
        except (ValueError, TypeError): continue

        rec = {"meta_num": meta_num}

        # ── Text fields ──────────────────────────────────────────────────
        for f in ["nom_pilar", "macrometa", "nom_sector", "nombre_programa",
                   "cod_bpin", "acciones_plan", "unidad_medida", "linea_base",
                   "indicador_meta", "secretaria", "descripcion_meta", "tipo_ponderado"]:
            rec[f] = clean_text(ws.cell(r, COL[f]).value)

        # DB has NOT NULL on secretaria — new metas without one yet get placeholder
        if not rec["secretaria"]:
            rec["secretaria"] = "POR DEFINIR"

        # ── Integer fields ───────────────────────────────────────────────
        for f in ["cod_dependencia", "num_pilar", "cod_sector", "cod_programa",
                   "anio_base", "codigo_producto"]:
            v = clean_num(ws.cell(r, COL[f]).value)
            rec[f] = int(v) if v is not None else None

        rec["meta_cuatrienio"] = clean_num(ws.cell(r, COL["meta_cuatrienio"]).value)

        # ── Per-year metas (NP → NULL) ───────────────────────────────────
        for yr in VALID_YEARS:
            pdm_val = ws.cell(r, COL[f"meta_pdm_{yr}"]).value
            fis_val = ws.cell(r, COL[f"meta_fisica_{yr}"]).value
            rec[f"meta_pdm_{yr}"] = None if is_np(pdm_val) else clean_num(pdm_val)
            rec[f"meta_fisica_{yr}"] = None if is_np(fis_val) else clean_num(fis_val)

        # ── Presupuesto per year (JSONB) ─────────────────────────────────
        for yr in VALID_YEARS:
            ta = clean_currency(ws.cell(r, COL[f"total_apropiacion_{yr}"]).value)
            imder_key = f"total_apropiacion_imder_{yr}"
            if imder_key in COL:
                ta_imder = clean_currency(ws.cell(r, COL[imder_key]).value)
                if ta_imder and ta_imder > 0:
                    ta = (ta or 0) + ta_imder
            rec[f"presupuesto_{yr}"] = {
                "total_apropiacion": ta,
                "neto_registros": clean_currency(ws.cell(r, COL[f"neto_registros_{yr}"]).value),
                "total_obligacion": clean_currency(ws.cell(r, COL[f"total_obligacion_{yr}"]).value),
                "pct_ejecucion_obligado": clean_pct(ws.cell(r, COL[f"pct_ejec_obligado_{yr}"]).value),
                "pct_ejecucion_pagos": clean_pct(ws.cell(r, COL[f"pct_ejec_pagos_{yr}"]).value),
                "presupuesto_comprometer": clean_currency(ws.cell(r, COL[f"presup_comprometer_{yr}"]).value),
            }

        # ── Eficiencia (as-is from Excel) ────────────────────────────────
        for yr in VALID_YEARS:
            rec[f"eficiencia_{yr}"] = clean_pct(ws.cell(r, COL[f"eficiencia_{yr}"]).value)

        # ── Ponderado avance per year = Avance Fisico año (as-is) ────────
        for yr in VALID_YEARS:
            rec[f"ponderado_avance_{yr}"] = clean_num(ws.cell(r, COL[f"avfis_{yr}"]).value)

        # ── Avance Fisico Cuatrienio (as-is from Excel) ─────────────────
        rec["avance_fisico"] = clean_num(ws.cell(r, COL["avance_fisico"]).value)
        rec["ponderado_cuatrienio"] = rec["avance_fisico"]

        # ── Cumplimiento Cuatrienio (as-is, Excel stores 0-1 scale) ─────
        cc = clean_num(ws.cell(r, COL["cumplimiento_cuatrienio"]).value)
        # DB stores cumplimiento_cuatrienio in 0-100 scale
        rec["cumplimiento_cuatrienio"] = cc * 100 if cc is not None and cc <= 1.0 else cc

        # ── Observaciones / Compromisos ──────────────────────────────────
        for yr in VALID_YEARS:
            rec[f"observaciones_{yr}"] = clean_text(ws.cell(r, COL[f"observaciones_{yr}"]).value)
            rec[f"compromisos_{yr}"] = clean_text(ws.cell(r, COL[f"compromisos_{yr}"]).value)

        records.append(rec)

    # Deduplicate by meta_num — keep last occurrence (most complete row)
    seen = {}
    for rec in records:
        seen[rec["meta_num"]] = rec
    records = list(seen.values())

    print(f"[2] Records parsed: {len(records)}")

    # ── Validation ──────────────────────────────────────────────────────────
    print(f"\n[3] VALIDATION:")
    for yr in VALID_YEARS:
        ef_vals = [rec[f"eficiencia_{yr}"] for rec in records if rec[f"eficiencia_{yr}"] is not None]
        avg_ef = sum(ef_vals) / len(ef_vals) * 100 if ef_vals else 0
        print(f"  Eficiencia {yr}: avg={avg_ef:.1f}% (n={len(ef_vals)})")

    avfs = [rec["avance_fisico"] for rec in records if rec["avance_fisico"] is not None]
    print(f"  Avance Fisico Cuatrienio: avg={sum(avfs)/len(avfs)*100:.1f}% (n={len(avfs)})")

    cumpls = [rec["cumplimiento_cuatrienio"] for rec in records if rec["cumplimiento_cuatrienio"] is not None]
    print(f"  Cumplimiento Cuatrienio: avg={sum(cumpls)/len(cumpls):.1f}% (n={len(cumpls)})")

    for yr in VALID_YEARS:
        total = sum((rec[f"presupuesto_{yr}"]["total_apropiacion"] or 0) for rec in records)
        print(f"  Total Apropiacion {yr}: ${total:,.0f}")

    secs = set(rec["secretaria"] for rec in records if rec["secretaria"])
    tipos = {}
    for rec in records:
        tp = rec["tipo_ponderado"] or "NULL"
        tipos[tp] = tipos.get(tp, 0) + 1
    print(f"  Secretarias: {len(secs)} | Tipos: {tipos}")

    # ── Generate SQL ────────────────────────────────────────────────────────
    print(f"\n[4] Generating SQL...")

    db_cols = [
        "meta_num", "cod_dependencia", "num_pilar", "nom_pilar", "macrometa",
        "cod_sector", "nom_sector", "cod_programa", "nombre_programa", "cod_bpin",
        "acciones_plan", "unidad_medida", "linea_base", "anio_base", "codigo_producto",
        "indicador_meta", "secretaria", "descripcion_meta", "meta_cuatrienio", "tipo_ponderado",
        "meta_pdm_2024", "meta_fisica_2024", "meta_pdm_2025", "meta_fisica_2025",
        "meta_pdm_2026", "meta_fisica_2026", "meta_pdm_2027", "meta_fisica_2027",
        "presupuesto_2024", "presupuesto_2025", "presupuesto_2026", "presupuesto_2027",
        "ponderado_avance_2024", "ponderado_avance_2025", "ponderado_avance_2026", "ponderado_avance_2027",
        "eficiencia_2024", "eficiencia_2025", "eficiencia_2026", "eficiencia_2027",
        "avance_fisico", "ponderado_cuatrienio", "cumplimiento_cuatrienio",
        "observaciones_2024", "compromisos_2024", "observaciones_2025", "compromisos_2025",
        "observaciones_2026", "compromisos_2026", "observaciones_2027", "compromisos_2027",
    ]
    text_cols = {
        "nom_pilar", "macrometa", "nom_sector", "nombre_programa", "cod_bpin",
        "acciones_plan", "unidad_medida", "linea_base", "indicador_meta",
        "secretaria", "descripcion_meta", "tipo_ponderado",
        "observaciones_2024", "compromisos_2024", "observaciones_2025", "compromisos_2025",
        "observaciones_2026", "compromisos_2026", "observaciones_2027", "compromisos_2027",
    }
    jsonb_cols = {"presupuesto_2024", "presupuesto_2025", "presupuesto_2026", "presupuesto_2027"}

    sql = []
    sql.append(f"-- Migracion 013: Recarga completa PDM desde '{INPUT_FILE}'")
    sql.append(f"-- Generado por pipeline_pdm_2026.py — valores tal cual del Excel, sin caps")
    sql.append(f"-- {len(records)} metas | Avg Avance Fisico: {sum(avfs)/len(avfs)*100:.1f}%")
    sql.append("")
    sql.append("BEGIN;")
    sql.append("TRUNCATE pdm_metas RESTART IDENTITY;")
    sql.append("")

    for rec in records:
        vals = []
        for c in db_cols:
            v = rec.get(c)
            if c in jsonb_cols: vals.append(sql_jsonb(v))
            elif c in text_cols: vals.append(sql_val(v, is_text=True))
            else: vals.append(sql_val(v))
        sql.append(f"INSERT INTO pdm_metas ({', '.join(db_cols)}) VALUES ({', '.join(vals)});")

    sql.append("")
    sql.append("COMMIT;")
    sql.append("")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(sql))
    print(f"[5] SQL generated: {OUTPUT_PATH}")
    print(f"    {len(records)} INSERTs")

    # ── Validation CSV ──────────────────────────────────────────────────────
    csv_path = Path(__file__).parent / "pdm_validacion_2026.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["meta_num","secretaria","tipo","MC",
                     "pdm24","fis24","ef24","avfis24",
                     "pdm25","fis25","ef25","avfis25",
                     "pdm26","fis26","ef26","avfis26",
                     "pdm27","fis27","ef27","avfis27",
                     "avance_fisico","cumplimiento",
                     "aprop24_M","aprop25_M","aprop26_M","aprop27_M"])
        for rec in records:
            row = [rec["meta_num"], rec["secretaria"], rec["tipo_ponderado"], rec["meta_cuatrienio"]]
            for yr in VALID_YEARS:
                ef = rec[f"eficiencia_{yr}"]
                pa = rec[f"ponderado_avance_{yr}"]
                row.extend([
                    rec[f"meta_pdm_{yr}"], rec[f"meta_fisica_{yr}"],
                    f"{ef*100:.1f}%" if ef is not None else "",
                    f"{pa*100:.1f}%" if pa is not None else "",
                ])
            af = rec["avance_fisico"]
            cc = rec["cumplimiento_cuatrienio"]
            row.append(f"{af*100:.1f}%" if af is not None else "")
            row.append(f"{cc:.1f}%" if cc is not None else "")
            for yr in VALID_YEARS:
                a = rec[f"presupuesto_{yr}"]["total_apropiacion"]
                row.append(f"{a/1e6:.1f}" if a else "0")
            w.writerow(row)
    print(f"[6] Validation CSV: {csv_path}")

    # ── Summary ─────────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  Total metas:        {len(records)}")
    print(f"  Acumulativo:        {tipos.get('Acumulativo', 0)}")
    print(f"  No acumulativo:     {tipos.get('No acumulativo', 0)}")
    print(f"  Secretarias:        {len(secs)}")
    print(f"  Avg Avance Fisico:  {sum(avfs)/len(avfs)*100:.1f}%")
    print(f"  Avg Cumplimiento:   {sum(cumpls)/len(cumpls):.1f}%")
    for yr in VALID_YEARS:
        vals = [rec[f"eficiencia_{yr}"] for rec in records if rec[f"eficiencia_{yr}"] is not None]
        print(f"  Avg Eficiencia {yr}: {sum(vals)/len(vals)*100:.1f}% (n={len(vals)})" if vals else f"  Avg Eficiencia {yr}: N/A")
    print(f"{'='*60}")
    print(f"\n  To apply:\n  cd alcaldia-geovisor && supabase db query --linked -f backend/src/db/migrations/{OUTPUT_SQL}")


if __name__ == "__main__":
    main()
